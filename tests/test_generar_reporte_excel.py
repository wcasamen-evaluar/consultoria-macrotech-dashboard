import importlib.util
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


RUTA_SCRIPT = (
    Path(__file__).resolve().parents[1]
    / "reporte excel"
    / "generar_reporte_excel.py"
)
SPEC = importlib.util.spec_from_file_location("generar_reporte_excel", RUTA_SCRIPT)
MODULO = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(MODULO)


class GenerarReporteExcelTest(unittest.TestCase):
    def test_consolida_evaluaciones_y_conserva_vacios(self):
        integrado = pd.DataFrame(
            [
                {
                    "colaborador": "Persona Completa",
                    "email_colaborador_360": "completa@example.com",
                    "evd_360": 91.234,
                    "potencial": 88.765,
                    "objetivos": 97.456,
                },
                {
                    "colaborador": "Solo Competencias",
                    "correo": "competencias@example.com",
                    "evd_360": pd.NA,
                    "potencial": 86.5,
                    "objetivos": pd.NA,
                },
                {
                    "colaborador": "Solo Objetivos",
                    "email_colaborador_obj": "objetivos@example.com",
                    "evd_360": pd.NA,
                    "potencial": pd.NA,
                    "objetivos": 100.0,
                },
            ]
        )

        reporte = MODULO.construir_reporte(integrado).set_index("Nombre")

        self.assertEqual(reporte.loc["Persona Completa", "desempeño"], 91.23)
        self.assertEqual(reporte.loc["Persona Completa", "competencias"], 88.77)
        self.assertEqual(reporte.loc["Persona Completa", "objetivos"], 97.46)
        self.assertTrue(pd.isna(reporte.loc["Solo Competencias", "desempeño"]))
        self.assertEqual(reporte.loc["Solo Competencias", "competencias"], 86.5)
        self.assertTrue(pd.isna(reporte.loc["Solo Competencias", "objetivos"]))
        self.assertEqual(
            reporte.loc["Solo Objetivos", "correo"], "objetivos@example.com"
        )

    def test_genera_cinco_columnas_y_celdas_vacias_reales(self):
        reporte = pd.DataFrame(
            [
                ["Persona Uno", "persona@example.com", 89.53, pd.NA, 95.0],
            ],
            columns=MODULO.COLUMNAS_SALIDA,
        )

        with tempfile.TemporaryDirectory() as carpeta:
            salida = Path(carpeta) / "reporte.xlsx"
            MODULO.guardar_excel(reporte, salida)
            libro = load_workbook(salida, data_only=True)
            hoja = libro["Reporte general"]

            self.assertEqual(
                [hoja.cell(1, columna).value for columna in range(1, 6)],
                MODULO.COLUMNAS_SALIDA,
            )
            self.assertEqual(hoja.max_column, 5)
            self.assertEqual(hoja["C2"].value, 89.53)
            self.assertIsNone(hoja["D2"].value)
            self.assertEqual(hoja["E2"].value, 95.0)


if __name__ == "__main__":
    unittest.main()
